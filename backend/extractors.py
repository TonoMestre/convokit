"""
ConvoKit — extracción de texto de documentos.

Extrae el contenido textual de cada formato admitido usando librerías locales.
Ninguna llamada externa: cero tokens de Claude consumidos en esta fase.

Paso 3: extracción de PDF (PyMuPDF).
Paso 4: se añade DOCX (python-docx) y XLSX (openpyxl).
"""

import io

import docx          # python-docx
import fitz          # PyMuPDF
import openpyxl


# Etiquetas que se muestran como cabecera en el contexto compuesto.
LABEL_HEADERS = {
    "bases_reguladoras":   "BASES REGULADORAS",
    "convocatoria":        "CONVOCATORIA DEL EJERCICIO",
    "plantilla_memoria":   "PLANTILLA DE MEMORIA / SOLICITUD",
    "resolucion_anterior": "RESOLUCIÓN DE EJERCICIO ANTERIOR",
    "anexo":               "ANEXO O DOCUMENTO COMPLEMENTARIO",
}


def extract_text(content: bytes, filename: str) -> str:
    """
    Detecta el formato por extensión y delega en el extractor correspondiente.
    Lanza ValueError con mensaje en español si el formato no es compatible o
    si el PDF no tiene texto extraíble (escaneado).
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "pdf":
        return _extract_pdf(content, filename)
    elif ext == "docx":
        return _extract_docx(content, filename)
    elif ext == "xlsx":
        return _extract_xlsx(content, filename)
    elif ext == "txt":
        return _extract_txt(content)
    else:
        raise ValueError(
            f"Formato no soportado: .{ext}. "
            "Sube el archivo en PDF, DOCX, XLSX o TXT."
        )


def build_context(documents: list[dict]) -> str:
    """
    Combina el texto extraído de todos los documentos en un único contexto
    compuesto, prefijando cada bloque con su etiqueta.

    Si el contexto total supera 150.000 palabras, trunca por orden de
    prioridad: bases_reguladoras → convocatoria → plantilla_memoria →
    resolucion_anterior → anexo.
    """
    PRIORITY = [
        "bases_reguladoras",
        "convocatoria",
        "plantilla_memoria",
        "resolucion_anterior",
        "anexo",
    ]
    MAX_WORDS = 150_000

    # Ordenar por prioridad para aplicar el límite correctamente.
    sorted_docs = sorted(
        documents,
        key=lambda d: PRIORITY.index(d["etiqueta"]) if d["etiqueta"] in PRIORITY else 99,
    )

    parts = []
    total_words = 0

    for doc in sorted_docs:
        header = LABEL_HEADERS.get(doc["etiqueta"], doc["etiqueta"].upper())
        block = f"=== {header} ===\n{doc['texto']}"
        words = len(block.split())

        if total_words + words > MAX_WORDS:
            # Incluir solo la parte que cabe.
            remaining = MAX_WORDS - total_words
            truncated = " ".join(block.split()[:remaining])
            parts.append(truncated + "\n[... documento truncado por límite de contexto ...]")
            break

        parts.append(block)
        total_words += words

    return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Extractores por formato
# ---------------------------------------------------------------------------

def _extract_pdf(content: bytes, filename: str) -> str:
    """
    Extrae texto de un PDF con PyMuPDF, página a página.
    Lanza ValueError si el PDF no tiene texto (escaneado).
    """
    try:
        doc = fitz.open(stream=content, filetype="pdf")
    except Exception:
        raise ValueError(f"No se pudo abrir el archivo '{filename}'. Comprueba que es un PDF válido.")

    pages_text = []
    for page in doc:
        pages_text.append(page.get_text())

    full_text = "\n".join(pages_text).strip()

    if not full_text:
        raise ValueError(
            "Este PDF parece estar escaneado y no contiene texto extraíble. "
            "Por favor, aporta la versión digital del documento."
        )

    return full_text


def _extract_docx(content: bytes, filename: str) -> str:
    """
    Extrae texto de un DOCX con python-docx.
    Mantiene el orden del documento original: párrafos en secuencia y tablas
    formateadas como texto tabulado (columnas separadas por " | ", filas por salto de línea).
    """
    try:
        document = docx.Document(io.BytesIO(content))
    except Exception:
        raise ValueError(f"No se pudo abrir el archivo '{filename}'. Comprueba que es un DOCX válido.")

    parts = []
    for block in document.element.body:
        tag = block.tag.split("}")[-1]  # "p" (párrafo) o "tbl" (tabla)

        if tag == "p":
            paragraph = docx.text.paragraph.Paragraph(block, document)
            text = paragraph.text.strip()
            if text:
                parts.append(text)

        elif tag == "tbl":
            table = docx.table.Table(block, document)
            for row in table.rows:
                cells = [cell.text.strip() for cell in row.cells]
                # Eliminar celdas duplicadas consecutivas (fusiones de celda).
                deduped = [cells[0]] + [c for i, c in enumerate(cells[1:], 1) if c != cells[i - 1]]
                row_text = " | ".join(deduped)
                if row_text.strip(" |"):
                    parts.append(row_text)

    full_text = "\n".join(parts).strip()
    if not full_text:
        raise ValueError(f"El archivo '{filename}' no contiene texto extraíble.")
    return full_text


def _extract_xlsx(content: bytes, filename: str) -> str:
    """
    Extrae texto de un XLSX con openpyxl, hoja a hoja.
    Cada hoja se precede de '=== HOJA: [nombre] ==='.
    Solo se incluyen celdas con contenido; se ignoran las vacías.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ValueError(f"No se pudo abrir el archivo '{filename}'. Comprueba que es un XLSX válido.")

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows():
            cells = [str(cell.value).strip() for cell in row if cell.value is not None]
            if cells:
                rows_text.append(" | ".join(cells))

        if rows_text:
            parts.append(f"=== HOJA: {sheet_name} ===")
            parts.extend(rows_text)

    full_text = "\n".join(parts).strip()
    if not full_text:
        raise ValueError(f"El archivo '{filename}' no contiene datos extraíbles.")
    return full_text


def _extract_txt(content: bytes) -> str:
    """
    Lee un fichero de texto plano en UTF-8 con fallback a latin-1.
    """
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("latin-1")
